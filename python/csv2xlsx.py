#!/usr/bin/env python3
"""
Lightweight CSV to XLSX Converter (no pandas required)
"""

import sys
import csv
import argparse
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def convert_csv_to_xlsx(input_file, output_file=None, delimiter=','):
    """Convert CSV to XLSX with proper UTF-8 handling"""
    input_path = Path(input_file)
    output_path = Path(output_file) if output_file else input_path.with_suffix('.xlsx')
    
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")
    
    print(f"Converting {input_path} to {output_path}...")
    
    # Read CSV data
    data = []
    encodings_to_try = ['utf-8-sig', 'utf-8', 'gbk', 'gb2312', 'big5']
    
    for encoding in encodings_to_try:
        try:
            with open(input_path, 'r', encoding=encoding, newline='') as csvfile:
                reader = csv.reader(csvfile, delimiter=delimiter)
                data = [row for row in reader if any(cell.strip() for cell in row)]
            print(f"Successfully read CSV with encoding: {encoding}")
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("Could not read CSV file with any supported encoding")
    
    if not data:
        raise ValueError("No data found in CSV file")
    
    # Create Excel workbook
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Data'
    
    # Write data to worksheet
    for row_num, row_data in enumerate(data, 1):
        for col_num, cell_value in enumerate(row_data, 1):
            worksheet.cell(row=row_num, column=col_num, value=cell_value)
    
    # Style header row
    if data:
        header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        header_font = Font(bold=True)
        
        for col_num in range(1, len(data[0]) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Save workbook
    workbook.save(output_path)
    
    print(f"Conversion completed! Output: {output_path}")
    print(f"Processed {len(data)} rows")
    
    return str(output_path)

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Convert CSV to XLSX')
    parser.add_argument('input_file', help='Input CSV file')
    parser.add_argument('output_file', nargs='?', help='Output XLSX file (optional)')
    parser.add_argument('--delimiter', '-d', default=',', help='CSV delimiter')
    
    args = parser.parse_args()
    
    try:
        convert_csv_to_xlsx(args.input_file, args.output_file, args.delimiter)
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)
